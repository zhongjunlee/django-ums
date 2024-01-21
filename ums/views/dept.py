from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt
from ums import models
from ums.utils.pagination import Pagination
from ums.utils.form import DepartmentForm


def dept_list(request):
    data_dict = {}
    search_data = request.GET.get('q', "")
    if search_data:
        data_dict["name__contains"] = search_data

    queryset = models.Department.objects.filter(**data_dict).order_by('-id')

    page_object = Pagination(request, queryset)

    form = DepartmentForm()

    context = {
        "form": form,
        "search_data": search_data,
        "queryset": page_object.page_queryset,  # 分完页的数据
        "page_string": page_object.html()  # 页码
    }

    return render(request, 'dept_list.html', context)


@csrf_exempt
def dept_add(request):
    form = DepartmentForm(data=request.POST)
    if form.is_valid():
        form.save()
        return JsonResponse({"status": True})
    return JsonResponse({'status': False, 'error': form.errors})


def dept_delete(request):
    uid = request.GET.get('uid')
    exists = models.Department.objects.filter(id=uid).exists()
    if not exists:
        return JsonResponse({'status': False, 'error': "删除失败，数据不存在"})

    models.Department.objects.filter(id=uid).delete()
    return JsonResponse({'status': True})


@csrf_exempt
def dept_edit(request):
    uid = request.GET.get('uid')
    row_object = models.Department.objects.filter(id=uid).first()
    if not row_object:
        return JsonResponse({'status': False, 'error': '数据不存在'})

    form = DepartmentForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return JsonResponse({'status': True})
    return JsonResponse({'status': False, 'error': form.errors})


def dept_detail(request):
    """ 根据ID获取订单详细 """
    uid = request.GET.get("uid")
    row_dict = models.Department.objects.filter(id=uid).values('name').first()
    if not row_dict:
        return JsonResponse({"status": False, 'error': "数据不存在"})

    # 从数据库中获取到一个对象 row_object
    result = {
        "status": True,
        "data": row_dict
    }
    return JsonResponse(result)


def dept_multi(request):
    """ 批量删除（Excel文件）"""
    from openpyxl import load_workbook

    # 1.获取用户上传的文件对象
    file_object = request.FILES.get("exc")

    # 2.对象传递给openpyxl，由openpyxl读取文件的内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    # 3.循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exists = models.Department.objects.filter(name=text).exists()
        if not exists:
            models.Department.objects.create(name=text)

    return redirect('/dept/list/')